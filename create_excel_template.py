import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_catalog_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catálogo Mikado Skincare"

    # Enforce grid lines visibility
    ws.views.sheetView[0].showGridLines = True

    # Headers definition as requested by user
    headers = [
        "Codigo",       # Código Único (SKU)
        "Marca",        # Marca del producto
        "Producto",     # Nombre completo del producto
        "Categoria",    # Categoria (Solares, Serums, Tonicos, Limpiadores, Hidratantes, Mascarillas)
        "Precio",       # Precio regular en Soles (S/)
        "Stock",        # Cantidad en inventario
        "Descuento",    # % Descuento o S/ Descuento (opcional)
        "Descripcion"   # Descripción corta
    ]

    ws.append(headers)

    # Initial K-Beauty sample catalog data
    sample_data = [
        ["MKD-001", "Beauty of Joseon", "Relief Sun: Rice + Probiotics SPF50+ PA++++", "Solares", 79.00, 25, 0.15, "Protector solar orgánico ligero con 30% extracto de arroz. No deja capa blanca."],
        ["MKD-002", "COSRX", "Advanced Snail 96 Mucin Power Essence 100ml", "Serums", 85.00, 18, 0.10, "Esencia con 96% mucina de caracol. Repara la barrera cutánea y aporta brillo Glass Skin."],
        ["MKD-003", "Anua", "Heartleaf 77% Soothing Toner 250ml", "Tonicos", 89.00, 30, 0.00, "Tónico calmante para pieles sensibles y propensas a rojeces o acné."],
        ["MKD-004", "Skin1004", "Madagascar Centella Ampoule 100ml", "Serums", 92.00, 12, 0.05, "Serum concentrado 100% Centella Asiática. Calma e hidrata intensamente."],
        ["MKD-005", "Laneige", "Lip Sleeping Mask EX [Berry] 20g", "Mascarillas", 75.00, 40, 0.10, "Mascarilla nocturna de labios de frutos rojos. Hidratación profunda 24h."],
        ["MKD-006", "Ma:nyo Factory", "Pure Cleansing Oil 200ml", "Limpiadores", 98.00, 15, 0.00, "Aceite limpiador hidrofílico con 14 aceites vegetales. Remueve todo el maquillaje."],
        ["MKD-007", "Round Lab", "Birch Juice Moisturizing Cream 80ml", "Hidratantes", 95.00, 20, 0.00, "Crema hidratante ligera con Sabia de Abedul e Hialurónico."],
        ["MKD-008", "COSRX", "Low pH Good Morning Gel Cleanser 150ml", "Limpiadores", 52.00, 50, 0.20, "Limpiador facial suave en gel con pH balanceado y BHA natural."]
    ]

    for row in sample_data:
        ws.append(row)

    # Styling Palette
    header_fill = PatternFill(start_color="D48C95", end_color="D48C95", fill_type="solid") # Rose Gold / Soft Pink
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    data_font = Font(name="Calibri", size=11)
    code_font = Font(name="Calibri", size=11, bold=True, color="2D2727")
    
    thin_border = Border(
        left=Side(style='thin', color='EFE8E1'),
        right=Side(style='thin', color='EFE8E1'),
        top=Side(style='thin', color='EFE8E1'),
        bottom=Side(style='thin', color='EFE8E1')
    )

    # Apply Header Styles
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Format Data Rows
    for row_idx in range(2, len(sample_data) + 2):
        # Code column
        ws.cell(row=row_idx, column=1).font = code_font
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")
        
        # Brand & Category
        ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal="center", vertical="center")
        
        # Price formatting (S/)
        price_cell = ws.cell(row=row_idx, column=5)
        price_cell.number_format = '"S/" #,##0.00'
        price_cell.alignment = Alignment(horizontal="right", vertical="center")

        # Stock formatting
        stock_cell = ws.cell(row=row_idx, column=6)
        stock_cell.number_format = '#,##0'
        stock_cell.alignment = Alignment(horizontal="right", vertical="center")

        # Discount formatting (%)
        disc_cell = ws.cell(row=row_idx, column=7)
        disc_cell.number_format = '0%'
        disc_cell.alignment = Alignment(horizontal="right", vertical="center")

        # Borders for all cells in row
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).border = thin_border

    # Adjust Column Widths automatically
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    ws.column_dimensions['C'].width = 45 # Product name
    ws.column_dimensions['H'].width = 50 # Description

    file_path = "catalogo_mikado.xlsx"
    wb.save(file_path)
    print(f"Template guardado exitosamente en: {file_path}")

if __name__ == "__main__":
    generate_catalog_excel()
